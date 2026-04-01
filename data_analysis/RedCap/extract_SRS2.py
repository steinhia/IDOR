from collections import Counter

import pandas as pd
import pdb
import numpy as np
data_folder="data/"
csv_folder="csv/"
import subprocess
from openpyxl import load_workbook
from extract_csv import extract_neuro_classification, extract_demographics

def question_col_names(test_type="escolar"):
    """
    return the names of the question columns
    """
    return [f"{i}" for i in range(1, 65 if test_type=="pre" else 66)]

def replace_nan(response,test_type,sex):
    """
    check the answer and if it is empty or nan, replaces it with the value that should be used instead (median of the responses in the population)
    :param response: response for the question
    :param test_type: escolar, pre-escolar, adulto a, adulto h
    :param sex: participant sex, F or M for escolar, 0 for other test types
    :return: the updated answer
    """
    ## TODO Not implemented yet
    return int(response) if (str(response).isdigit() and int(response) > 0) else 1

def replace_nans(df):
    """
    replace all nan values withthe median of the response forthis question of the test category
    :param df: dataframe with SRS2 results
    :return: cleaned dataframe
    """
    def replace_one_row(row):
        cols = question_col_names(row["test_type"])
        for col in cols:
            val = row[col]
            if pd.isna(val) or str(val).strip() == "":
                question = int(col)  # adapte si "Q1"
                row[col] = replace_nan(question, row["test_type"], row["sex"] )
        return row
    df = df.apply(lambda row: replace_one_row(row), axis=1)
    return df

def calculate_SRS2_score_id(id_subject,responses,test_type,sex):
    """
    uses the xls file to calculate the SRS2 scores for some subject
    :param id_subject: string. ID of the subject from which the function calculates the score
    :param responses: list of responses
    :param test_type: pre, escolar, adulto a, adulto h
    :param sex : string, sex of the participant.
    :return: dataframe with the results of the participant
    """
    # load file
    wb = load_workbook(data_folder + "SRS2_correction.xlsx")
    corr = wb["Correção"]

    # 2. Write results
    corr['B1'] = test_type
    corr['B2'] = int(0) if sex =='0' else sex
    for idx,i in enumerate(range(5, 70)):
        # to handle cases where non responses were marked as ' ', x, -1 ...
        rep=responses[idx]
        corr['C' + str(i)] = replace_nan(rep,test_type, sex)

    # 3. save temp file
    temp_path = csv_folder+"temp_input.xlsx"
    wb.save(temp_path)

    # 4. Calculation with LibreOffice headless
    subprocess.run(["libreoffice", "--headless", "--calc", "--convert-to", "xlsx", "--outdir", ".", temp_path], check=True)

    # 5. Read calculated values
    keys = ["TOTAL", "PERCEPÇÃO SOCIAL", "COGNIÇÃO SOCIAL", "COMUNICAÇÃO SOCIAL", "MOTIVAÇÃO SOCIAL", "PADRÕES RESTRITOS E REPETITIVOS", "COMUNICAÇÃO E INTERAÇÃO SOCIAL"]
    wb2 = load_workbook("temp_input.xlsx", data_only=True)  # <-- data_only=True lit les valeurs, pas les formules
    ws2 = wb2["Correção"]
    values = [ws2[f"I{row}"].value for row in range(5, 12)]
    return pd.DataFrame({ "subj_id": [id_subject] * len(keys), "key": keys, "value": values })

def calculate_SRS2_scores(df, repeat_calculation=True, HCP=False):
    """
    calculate the results (subscales) for all participants
    :param df: dataframe with the data of all subjects
    :param repeat_calculation: if True, make the calculation from the xls again (slow)
    :param HCP: if True, calculation for the controls from the HCP database
    :return: dataframe with the results for all participants
    """
    name="SRS2_results"+("_HCP" if HCP else "")+".csv"
    if repeat_calculation:
        all_results = []
        n_row=len(df)
        for x, row in df.iterrows():
            print(f"extract result for subject {x}/{n_row}")
            id_subject = row["subj_id"]
            test_type = row["test_type"]
            sex=row["sex"] if test_type=="escolar" else 0
            responses = [replace_nan(i,test_type, sex)  for i in row.iloc[3:68].tolist()]
            all_results.append(calculate_SRS2_score_id(id_subject, responses,test_type,sex))
        final_df = pd.concat(all_results, ignore_index=True)
        final_df.to_csv(csv_folder + name[:-4]+"_long.csv" , index=False)
        df_wide = final_df.pivot(index='subj_id', columns='key', values='value').reset_index()
        df_wide.to_csv(csv_folder + name, index=False)
    else:
        df_wide=pd.read_csv(csv_folder + name)
    return df_wide



#resultats = calculate_SRS2_scores()

def extract_SRS2():
    """
    extracts the relevant structure from the Redcap csv.
    The difficulty is that the 3 types of tests are not grouped into the same Q1...Q64, but instead columns are truplicated.
    :return: the formatted dataframe
    """
    df=pd.read_csv(data_folder+"SRS2.csv")
    dem=extract_demographics()
    df = df.merge(dem[['subj_id', 'sex']], on='subj_id', how='left')
    # sex on 3d position
    cols = df.columns.tolist()
    cols.insert(2, cols.pop(cols.index('sex')))
    df = df[cols]
    # numerotation problem in the es test (missing column 2 -> all other columns are shifted from 1)
    df = df.rename(columns={ "srs2_botao": "test_type", **{f"srs2_es_{i}": f"srs2_es_{i - 1}" for i in range(2, 68)} })
    df = df[df['test_type'].notna()].copy()
    tests = {1: "pre", 2: "escolar", 3: "adulto a", 4: "adulto h"}
    df['test_type']=df['test_type'].map(tests)
    test_type_Redcap={'pre':'ps','escolar':'es','adulto h':'ah','adulto a':'aa'}
    # regroup different test types under same columns (response numero)
    def extract_answers(row):
        test_type = row['test_type']  # ps / es / aa / ah
        # pre escolar only 64 questions
        question_col=question_col_names(test_type)
        cols = [f"srs2_{test_type_Redcap[test_type]}_{i}" for i in question_col]
        return pd.Series(row[cols].values, index=[f"{i}" for i in range(1, 65 if test_type=="pre" else 66)])
    df = df[['subj_id', 'test_type','sex']].join(df.apply(extract_answers, axis=1))
    # keep last test
    df = df.groupby('subj_id').last().reset_index()
    df=replace_nans(df)
    return df

def extract_srs2_HCP(repeat_calculation=True):
    """
    formats the results from the HCP database and calculates the SRS2 results
    :return: dataframe with the results
    """
    # reading all csv results
    df=pd.read_csv(data_folder+"srs02_HCP.txt",sep="\t")
    # selecting only the matching participants
    matching=pd.read_csv(csv_folder+"matching.csv")
    control_list = list(matching.subjectkey.values)
    controls = df[df.subjectkey.isin(control_list)]
    # modifying the dataframe to correspond to the IDOR dataset
    controls = controls.rename(columns={"src_subject_id": "subj_id"})
    controls["test_type"] = "escolar"
    parent_cols = [col for col in controls.columns if col.startswith("parentreport_")]
    rename_dict = {col: col.split("_")[1] for col in parent_cols}
    controls = controls.rename(columns=rename_dict)
    question_cols = [rename_dict[col] for col in parent_cols]
    final_cols = ["subj_id", "test_type","sex"] + question_cols
    controls = controls[final_cols]
    return controls
    # calculating the SRS2 score

    pdb.set_trace()

def extract_id_list(final_df):
    """
    extracts the list of participants to use elsewhere
    :param final_df: dataframe with all data
    :return: list of participants
    """
    l=list(set(final_df.subj_id.values))
    l.sort()
    print("lista de pacientes",l)
    return l


def extract_id_list_with_demographic(final_df):
    """
    calculates the list of all participants and demographic info used to do the matching
    :param final_df: SRS2 dataframe with all data
    :return: df with list of all participants and demographic info used to do the matching
    """
    l = extract_id_list(final_df)
    dem = extract_demographics()
    id_dem_list = dem[dem['subj_id'].isin(l)]
    id_dem_list.to_csv(csv_folder + "SRS2_Demographic.csv")


def first_analysis_SRS2(df):
    """
    first analyses that can be made from behavioral results and neuro classification
    :param df: SRS2 results
    """
    #neuro classification
    df_neuro = extract_neuro_classification()
    df = df.merge(df_neuro, on='subj_id', how='left')

    df["neuro_classification"] = df["neuro_classification"].fillna("control")
    df["isolado_classification"] = df["isolado_classification"].fillna("X")

    id_vars = ["subj_id", "group", "neuro_classification", "isolado_classification"]

    df = df.melt(
        id_vars=id_vars,
        var_name="test",
        value_name="score"
    )

    # Number of subjects of each group
    print("number of subjects of each group",df.groupby('subj_id').first().groupby('neuro_classification').count()['group'])

    # mean score
    print("mean score by scale \n",df.groupby('neuro_classification').agg({'score': 'mean'}))
    print("mean score by scale and Neuro classification\n ",df.groupby(['test','neuro_classification']).agg({'score': 'mean'}).unstack())

    df['above_60'] = df['score'] > 50
    print("proportion above threshold \n", df.groupby('test')['above_60'].mean())
    print("proportion above threshold \n ", df.groupby(['test','neuro_classification'])['above_60'].mean().unstack())


if __name__ == "__main__":
    HCP_df=extract_srs2_HCP()
    HCP_res=calculate_SRS2_scores(HCP_df, repeat_calculation=False, HCP=True)
    HCP_res['group']='controle'
    IDOR_df=extract_SRS2()
    IDOR_res=calculate_SRS2_scores(IDOR_df, repeat_calculation=False)
    IDOR_res['group']='paciente'
    id_list=extract_id_list_with_demographic(IDOR_res)
    res_total=pd.concat((HCP_res,IDOR_res))
    first_analysis_SRS2(res_total)
    pdb.set_trace()

